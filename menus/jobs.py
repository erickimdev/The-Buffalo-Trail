import pygame
from openpyxl import *
from random import *
from text import *
from button import *

class Jobs:
    def __init__(self, game):
        self.game = game

        # parse excel information
        self.jobs_ws = load_workbook(filename='assets/jobs.xlsx')['jobs']
        row = self.jobs_ws[randint(2, self.jobs_ws.max_row)]

        # instantiate headers
        self.job_title = Text(row[1].value, self.game.width / 2, 100, 70, WHITE)
        self.job_description = Text(row[2].value, self.game.width / 2, 200, 25, WHITE)
        self.percentage = Text("", 1120, 120, 30, WHITE)
        self.percentage.text = str(1 / (self.jobs_ws.max_row - 1) * 100) + "%"

        x_offset = 100
        y_offset = 300
        # instantiate SOLUTION_1 button
        self.solution1_button = Button(x_offset, y_offset+80, 200, 70, LIGHT_GRAY, False, self.game)
        self.solution1_button_text = Text('Solution 1', x_offset+100, y_offset+115, 25, BLACK)
        # instantiate SOLUTION_2 button
        self.solution2_button = Button(x_offset, y_offset+180, 200, 70, LIGHT_GRAY, False, self.game)
        self.solution2_button_text = Text('Solution 2', x_offset+100, y_offset+217, 25, BLACK)

        # instantiate BACK button
        self.back_button = Button(950, 640, 200, 80, LIGHT_GRAY, False, self.game)
        self.back_button_text = Text('Back', 1050, 683, 30, BLACK)

    def draw_screen(self):
        # draw header texts
        self.job_title.draw(self.game.screen)
        self.job_description.draw(self.game.screen)
        self.percentage.draw(self.game.screen)

        # draw SOLUTION_1 button
        self.solution1_button.draw(self.game.screen)
        self.solution1_button_text.draw(self.game.screen)
        # draw SOLUTION_2 button
        self.solution2_button.draw(self.game.screen)
        self.solution2_button_text.draw(self.game.screen)

        # draw BACK button
        self.back_button.draw(self.game.screen)
        self.back_button_text.draw(self.game.screen)

    def catch_actions(self, event, mx, my):
        # catch SOLUTION_1 button clicks
        self.solution1_button.change_menu(event, mx, my, "pitstop")
        # catch SOLUTION_2 button clicks
        self.solution2_button.change_menu(event, mx, my, "pitstop")

        # catch BACK button clicks
        self.back_button.change_menu(event, mx, my, "pitstop")
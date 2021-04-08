import pygame
from openpyxl import *
from random import *
from text import *
from button import *

minor_events_db = load_workbook(filename='assets/Book1.xlsx')
minor_events_ws = minor_events_db['minor_events']


class Popup:
    def __init__(self, game):
        self.game = game
        self.minor = False
        # instantiate RESUME button
        self.resume_button = Button(900, 500, 200, 70, LIGHT_GRAY, False, self.game)
        self.resume_text = Text("Ok", 1000, 535, 50, BLACK)
        self.event_name = Text("Dummy Test 1", 400, 150, 50, BLACK)
        self.event_description = Text("Dummy Description", 300, 220, 14, BLACK)
        self.percentage = Text("Dummy %", 900, 150, 20, BLACK)

        rand_row = randint(2, minor_events_ws.max_row)
        row = minor_events_ws[rand_row]
        self.event_name.text = row[1].value
        self.event_description.text = row[2].value
        self.percentage.text = str(1 / (minor_events_ws.max_row - 1) * 100) + "%"

    def draw_screen(self):
        pygame.draw.rect(self.game.screen, WHITE, (150, 50, 1000, 600))
        # RESUME button
        self.resume_button.draw(self.game.screen)
        self.resume_text.draw(self.game.screen)

        self.event_name.draw(self.game.screen)
        self.event_description.draw(self.game.screen)
        self.percentage.draw(self.game.screen)

    def currently_in_play(self):
        pygame.time.set_timer(3, 20000)

    def not_in_play(self):
        pygame.time.set_timer(3, 0)

    def catch_actions(self, event, mx, my):
        # press esc key to UNPAUSE
        if event.type == pygame.MOUSEBUTTONDOWN:
            self.minor = False




import pygame
from openpyxl import *
from random import *
from text import *
from button import *

jobs_db = load_workbook(filename='assets/Book1.xlsx')
jobs_ws = jobs_db['jobs']


class JobsPopUp:
    def __init__(self, game):
        self.game = game
        self.minor = False
        # instantiate RESUME button
        self.event_name = Text("Dummy Test 1", 400, 150, 50, BLACK)
        self.event_description = Text("Dummy Description", 400, 220, 14, BLACK)
        self.percentage = Text("Dummy %", 900, 150, 20, BLACK)

        self.solution1_button = Button(340, 400, 150, 50, GRAY, False, self.game)
        self.solution1_button_text = Text("Solution 1", 420, 420, 20, BLACK)
        self.solution2_button = Button(340, 500, 150, 50, GRAY, False, self.game)
        self.solution2_button_text = Text("Solution 2", 420, 520, 20, BLACK)

        rand_row = randint(2, jobs_ws.max_row)
        row = jobs_ws[rand_row]
        self.event_name.text = row[1].value
        self.event_description.text = row[2].value
        self.percentage.text = str(1 / (jobs_ws.max_row - 1) * 100) + "%"

    def draw_screen(self):
        pygame.draw.rect(self.game.screen, WHITE, (150, 50, 1000, 600))
        self.event_name.draw(self.game.screen)
        self.event_description.draw(self.game.screen)
        self.percentage.draw(self.game.screen)

        self.solution1_button.draw(self.game.screen)
        self.solution1_button_text.draw(self.game.screen)
        self.solution2_button.draw(self.game.screen)
        self.solution2_button_text.draw(self.game.screen)

    def catch_actions(self, event, mx, my):

        self.solution1_button.change_menu(event, mx, my, "play")

        self.solution2_button.change_menu(event, mx, my, "play")



